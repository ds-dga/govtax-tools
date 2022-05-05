from openpyxl import load_workbook
from datetime import datetime
import re
import os
from db import Database

MONTH_CONV = {
    "มค": 1,
    "กพ": 2,
    "มีค": 3,
    "เมย": 4,
    "พค": 5,
    "มิย": 6,
    "กค": 7,
    "สค": 8,
    "กย": 9,
    "ตค": 10,
    "พย": 11,
    "ธค": 12,
}

init = True


def compatibility_sake(txt):
    if not txt:
        return txt
    # replace อื่น ๆ -> อื่นๆ since govspending's DB does this
    txt = txt.replace("-", "")
    txt = txt.replace("อื่น ๆ", "อื่นๆ")
    txt = "ภาษีสุราฯ" if txt == "ภาษีสุรา" else txt
    return txt


def extract_mo(val):
    """there are 2 possibilities so far
    1.  ต.ค. 60 	 พ.ย. 60 	 ธ.ค. 60 	 ม.ค. 61
    2. Oct-63	Nov-63	Dec-63	Jan-64	Feb-64	Mar-64

    First one is just a text, but unfortunately
    later turns out to be python datetime --- 63 is correct, but it's 19-fucking-63!

    What we want?
    we need b.e.
    return (YYYY, MM) < str, int doesn't matter since it will be `int` in DB anyway
    """
    if isinstance(val, datetime):
        yr = f"{val.year}"[2:]
        return [f"25{yr}", val.month]

    cmp = r"(.*?)\s+(\d\d)"
    res = re.match(cmp, val)
    if res:
        mo, yr = res.groups()
        mo = mo.replace(".", "")
        yr = f"25{yr}"
        if mo in MONTH_CONV:
            return [yr, MONTH_CONV[mo]]
    return val


def get_months(sheet, row_ind):
    selected_row = sheet[row_ind]
    mos = [
        (*extract_mo(i.value), i.column, i.column_letter)
        for i in selected_row
        if i.value
    ]
    return mos


def get_income(sheet):
    """It will focus on finding department-section first; then getting available data later"""
    end = "รวมรายได้จัดเก็บ"
    departments = ["กรมสรรพากร", "กรมสรรพสามิต", "กรมศุลกากร", "หน่วยงานอื่น"]
    junks = [
        "รวม 3 กรม",
    ]
    start_row = None
    active_dept = None
    months = []
    results = []  # dept, section, yyyy-mm, value

    for i in sheet.iter_rows():
        val = i[0].value
        row_ind = i[0].row
        val = val.strip() if val else val

        if val in junks:
            continue
        # replace อื่น ๆ -> อื่นๆ since govspending's DB does this
        val = compatibility_sake(val)

        if not months and start_row:
            months = get_months(sheet, start_row - 1)
            # print('>> get months -->', months)

        if val == end:
            return results, months

        if val in departments:
            # print(val)
            if active_dept is None:
                start_row = i[0].row
            active_dept = val

        if active_dept and val != active_dept:
            # print(f"{active_dept}: {val}, row#{row_ind} mo#{len(months)} ")
            data = []
            for yyyy, mm, col_ind, col_name in months:
                data.append([yyyy, mm, sheet[f"{col_name}{row_ind}"].value])

            results.append([active_dept, val, data])
            # print(f'     {i[0]}, {i[0].value}')
            # print(f"     {start_row}")
    return results, months


def get_deduction(sheet, months):
    skips = ["รวมรายได้สุทธิ", "จัดสรรให้ อปท. ตาม พ.ร.บ. กำหนดแผนฯ"]
    end = [
        "สัดส่วนการคืนภาษีมูลค่าเพิ่ม",
    ]
    start = "หัก"
    started = False
    active_cat = None
    active_cat_data = []
    main_cat = True
    has_sub_data = False
    results = []

    for i in sheet.iter_rows():
        val = i[0].value
        row_ind = i[0].row
        val = val.strip() if val else val
        if not val:
            continue

        # replace อื่น ๆ -> อื่นๆ since govspending's DB does this
        val = compatibility_sake(val).strip()

        if not started and val == start:
            started = True

        if not started:
            continue

        if val in skips:
            continue

        if val in end:
            if not has_sub_data:
                results.append([active_cat, "", active_cat_data])
            return results

        cat_cmp = r"([\d\.]+)\s+?(.*)"
        sub_cmp = r"([-]+)\s+?(.*)"
        main_cat = re.match(cat_cmp, val)
        sub_cat = re.match(sub_cmp, val)

        if main_cat:
            val = main_cat.group(2)
        elif sub_cat:
            val = sub_cat.group(2).strip()

        """for god's sake Metamedia decided that "รวมรายได้สุทธิหลังหักจัดสรร" is
        one of deduction_type and change the name to รายได้สุทธิหลังหักจัดสรร
        """
        if val == "รวมรายได้สุทธิหลังหักจัดสรร":
            main_cat = True
            val = "รายได้สุทธิหลังหักจัดสรร"

        if main_cat:
            # print(val)
            if active_cat != val and active_cat:
                # enter new category
                # print("  > new cat")
                if not has_sub_data:
                    # print("  >> saved prev cat data")
                    results.append([active_cat, "", active_cat_data])

            active_cat = val
            has_sub_data = False
            active_cat_data = []
            for yyyy, mm, col_ind, col_name in months:
                active_cat_data.append([yyyy, mm, sheet[f"{col_name}{row_ind}"].value])
        elif active_cat:
            data = []
            for yyyy, mm, col_ind, col_name in months:
                data.append([yyyy, mm, sheet[f"{col_name}{row_ind}"].value])
            # print(' > ', data)
            results.append([active_cat, val, data])
            has_sub_data = True

    return results


def db_recorder(db, dataType, dept, section, yyyy, mm, total):
    budget_year = int(yyyy)
    if mm >= 10:
        budget_year += 1
    if dataType == "income":
        dept_id = db.get_or_create_revenue_dept(dept)
        revenue_type_id = db.get_or_create_revenue_type(section)
        # print(f"[recorder] {dataType} dID={dept_id} rID={revenue_type_id}")
        _id = db.insert_revenue_income(
            dept_id, revenue_type_id, budget_year, yyyy, mm, total
        )
        return _id
    elif dataType == "deduction":
        rd_type_name = dept
        if section:
            rd_type_name = f"{dept}: {section}"
        deduct_type_id = db.get_or_create_revenue_deduct_type(rd_type_name)
        # print(f"[recorder] {dataType} rID={deduct_type_id}")
        _id = db.insert_revenue_deduct(deduct_type_id, budget_year, yyyy, mm, total)
        return _id
    return False


def handle_file(fpath, verbose, db):
    global init
    init = True
    message = []
    message.append(f"[revenue] source path: {fpath}")
    wb = load_workbook(fpath)
    structured_data = []
    for sht_name in wb.sheetnames:
        sht = wb[sht_name]
        results, months = get_income(sht)
        for dept, section, data in results:
            if dept == "กรมสรรพสามิต" and section == "ภาษีอื่นๆ":
                # skip this since it's repetitive data
                continue
            for yyyy, mm, total in data:
                if not total:
                    continue
                if verbose:
                    print(f'[{yyyy}-{mm}] {f"{dept}-{section}".ljust(45)} {total:10f}')
                if db:
                    db_recorder(db, "income", dept, section, yyyy, mm, total)
                structured_data.append(
                    [
                        dept,
                        section,
                        yyyy,
                        mm,
                        total,
                    ]
                )
        results = get_deduction(sht, months)
        for dept, section, data in results:
            for yyyy, mm, total in data:
                if not total:
                    continue
                if verbose:
                    print(f'[{yyyy}-{mm}] {f"{dept}-{section}".ljust(45)} {total:10f}')
                if db:
                    db_recorder(db, "deduction", dept, section, yyyy, mm, total)
                structured_data.append(
                    [
                        dept,
                        section,
                        yyyy,
                        mm,
                        total,
                    ]
                )
    if verbose:
        latest = get_latest_month(structured_data)
        message.append(f"[revenue] Latest record on {latest.isoformat()}")
        message.append(f"[revenue] Total #{len(structured_data)}")
        # n = len(message) + 1
        # if not init:
        #     print(("\033[F\033[K") * n)
        print("\n".join(message))

    if len(structured_data) > 0 and db:
        latest = get_latest_month(structured_data)
        result_id = db.insert_data_source_update("revenue", "", latest.isoformat())
        result_msg = result_id if result_id else '--dup--'
        print(f"[revenue] data_source_update result ID #{result_msg}")

    init = False


def get_latest_month(items):
    yyyy, mm = 0, 0
    for [_, _, yr, mo, _] in items:
        if int(yr) > yyyy:
            yyyy = int(yr)
            mm = int(mo)
        elif int(yr) == yyyy and int(mo) > mm:
            mm = int(mo)
    # make sure date is correct on GMT+7 too
    return datetime(yyyy - 543, mm, 1, 14, 0, 0)


def handle_dir(fpath, verbose, db):
    for root, _, fs in os.walk(fpath):
        for f in fs:
            fp = os.path.join(root, f)
            if f.find(".xls") < 0 or f.find("~$") > -1:
                continue
            handle_file(fp, verbose, db)


def handle_source(path, **kw):
    verbose = save2db = False
    if "args" in kw:
        verbose = kw["args"].print
        save2db = kw["args"].db

    db = None
    if save2db:
        db = Database()

    if os.path.isdir(path):
        print(f"[revenue] {path} is directory")
        handle_dir(path, verbose, db)
    else:
        handle_file(path, verbose, db)
